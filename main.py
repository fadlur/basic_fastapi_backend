import os
from fastapi import FastAPI, Form, File, UploadFile, HTTPException, Depends
from fastapi import Request
from fastapi.responses import JSONResponse
from pydantic import BaseModel
from typing import Optional, Literal
from sqlalchemy.orm import Session
from database import Base, engine, SessionLocal
from models import FileMetadata
from openpyxl import load_workbook
import shutil

app = FastAPI()

# Buat database dan table
Base.metadata.create_all(bind=engine)

# Dependency DB Session
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# Folder lokasi penyimpanan
UPLOAD_DIRS = {
    "excel": "uploads/excels",
    "image": "uploads/images",
    "video": "uploads/videos",
}

os.makedirs("uploads", exist_ok=True)
for folder in UPLOAD_DIRS.values(): # karena kita akan mengambil value dari dict UPLOAD_DIRS
    os.makedirs(folder, exist_ok=True)

# Batasan ukuran file yang diupload
MAX_FILE_SIZE_MB = 10

# --------------------------
# GET Request
# --------------------------
@app.get("/items/{item_id}")
def read_item(item_id: int, q: Optional[str] = None):
    return {"method": "GET", "item_id": item_id, "query": q}

# --------------------------
# POST Request - JSON Body
# --------------------------
class Item(BaseModel):
    name: str
    description: Optional[str] = None
    price: float

@app.post("/items/json")
def create_item_json(item: Item):
    return {"method": "POST", "type": "application/json", "data": item.dict()}

# --------------------------
# POST Request - Form Data
# --------------------------
@app.post("/items/form")
def create_item_form(name: str = Form(...), price: float = Form(...)):
    return {"method": "POST", "type": "form-data", "name": name, "price": price}

# --------------------------
# POST Request - x-www-form-urlencoded
# --------------------------
@app.post("/items/xform")
async def create_item_xform(request: Request):
    form_data = await request.form()
    return {"method": "POST", "type": "x-www-form-urlencoded", "data": dict(form_data)}

# --------------------------
# POST Request - File Upload
# --------------------------
@app.post("/uploadfile/")
async def upload_file(file: UploadFile = File(...)):
    content = await file.read()
    return {
        "method": "POST",
        "type": "multipart/form-data",
        "filename": file.filename,
        "content_type": file.content_type,
        "size": len(content)
    }

# --------------------------
# PUT / PATCH Request
# --------------------------
@app.put("/items/{item_id}")
@app.patch("/items/{item_id}")
def update_item(item_id: int, item: Item):
    return {"method": "PUT/PATCH", "item_id": item_id, "updated_data": item.dict()}

# --------------------------
# DELETE Request
# --------------------------
@app.delete("/items/{item_id}")
def delete_item(item_id: int):
    return {"method": "DELETE", "item_id": item_id, "status": "deleted"}

# ---------------------------
# Tambahkan fungsi upload
# ---------------------------

# Helper untuk cek ukuran file
async def validate_file_size(file: UploadFile):
    content = await file.read()
    size_mb = len(content) / (1024 * 1024)
    await file.seek(0) # Reset pointer setelah dibaca
    if size_mb > MAX_FILE_SIZE_MB:
        raise HTTPException(status_code = 400, detail = f"File size exceeds {MAX_FILE_SIZE_MB} MB limit.")
    return content

# ---------------------------
# Fungsi upload excel
# ---------------------------
# @app.post("/upload/excel")
# async def upload_excel(file: UploadFile = File(...)):
#     # validasi extensi file excel (.xlsx sama .xls)
#     if not file.filename.endswith((".xlsx", ".xls")):
#         raise HTTPException(status_code=400, detail = "Only Excel files are allowed")
    
#     content = await validate_file_size(file)

#     # validasi content_type
#     if file.content_type not in [
#         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#         "application/vnd.ms-excel"
#     ]:
#         raise HTTPException(status_code=400, detail="Invalid Excel MIME type.")
    
#     return {"filename": file.filename, "type": "excel", "size_kb": f"{round(len(content) / (1024 * 1024), 2)} MB"}


# ---------------------------
# Fungsi upload image
# ---------------------------
# @app.post("/upload/image")
# async def upload_image(file: UploadFile = File(...)):
#     if not file.filename.lower().endswith((".png", ".jpg", ".jpeg")):
#         raise HTTPException(status_code=400, detail = "Only image files (jpg, png)")
    
#     content = await validate_file_size(file)

#     if file.content_type not in ["image/png", "image/jpeg"]:
#         raise HTTPException(status_code=400, detail="Invalid image MIME type.")
    
#     return {"filename": file.filename, "type": "image", "size_kb": f"{round(len(content) / (1024 * 1024), 2)} MB"}

# ---------------------------
# Fungsi upload video
# ---------------------------
# @app.post("/upload/video")
# async def upload_video(file: UploadFile = File(...)):
#     if not file.filename.lower().endswith((".mp4", ".avi", ".mov", ".mkv")):
#         raise HTTPException(status_code=400, detail="Only video files are allowed.")
    
#     content = await validate_file_size(file)

#     if not file.content_type.startswith("video/"):
#         raise HTTPException(status_code=400, detail="Invalid video MIME type.")
    
#     return {"filename": file.filename, "type": "video", "size_mb": f"{round(len(content) / (1024 * 1024), 2)} MB"}

# ---------------------------
# Refactor function upload
# ---------------------------

# Validasi dan simpan file
async def handle_upload(file: UploadFile, file_type: Literal["excel", "image", "video"], db: Session):
    ext_map = {
        "excel": ([".xlsx", ".xls"], ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet","application/vnd.ms-excel"]),
        "image": ([".png", ".jpg", ".jpeg"], ["image/png", "image/jpeg"]),
        "video": ([".mp4", ".avi", ".mov", ".mkv"], ["video/mp4", "video/x-msvideo", "video/quicktime", "video/x-matroska"])
    }

    allowed_exts, allowed_mimes = ext_map[file_type]
    filename = file.filename.lower()
    ext = os.path.splitext(filename)[1]

    if ext not in allowed_exts:
        raise HTTPException(status_code=400, detail=f"File extension {ext} not allowed for type {file_type}")
    
    if file.content_type not in allowed_mimes and not file.content_type.startswith("video/"):
        raise HTTPException(status_code=400, detail=f"MIME type {file.content_type} is invalid for {file_type}")
    
    content = await file.read()
    size_kb = len(content) / 1024
    await file.seek(0)

    # Simpan file ke folder
    save_path = os.path.join(UPLOAD_DIRS[file_type], filename)
    with open(save_path, "wb") as out_file: # context manager open file dalam mode biner (wb - write biner), out_file object file tujuan
        shutil.copyfileobj(file.file, out_file) # copy file.file ke out_file

    # Simpan metadata ke database
    metadata = FileMetadata(
        filename=filename,
        file_type=file_type,
        mime_type=file.content_type,
        size_kb=round(size_kb, 2),
        saved_path=save_path
    )

    db.add(metadata)
    db.commit()
    db.refresh(metadata)

    response = {
        "id": metadata.id,
        "filename": filename,
        "type": file_type,
        "mime": file.content_type,
        "size_kb": metadata.size_kb,
        "saved_to": save_path
    }

    # Jika Excel, tampilkan isi sheet pertama
    if file_type == "excel":
        wb = load_workbook(save_path)
        ws = wb.active
        data = [[cell.value for cell in row] for row in ws.iter_rows(min_row=1, max_row=5)]
        response["preview_data"] = data

    return response

# handle upload file jadi 1 function
@app.post("/upload/{file_type}", summary="Upload Excel/Image/Video with Validation")
async def upload_file(
    file_type: Literal["excel", "image", "video"],
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    return await handle_upload(file, file_type, db)