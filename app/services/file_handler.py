import os
from fastapi import UploadFile, HTTPException
import shutil
from models.file_metadata import FileMetadata
from openpyxl import load_workbook
from typing import Literal
from sqlalchemy.orm import Session

# Folder lokasi penyimpanan
UPLOAD_DIRS = {
    "excel": "app/uploads/excels",
    "image": "app/uploads/images",
    "video": "app/uploads/videos",
}
# Validasi dan simpan file
async def handle_upload(file: UploadFile, file_type: Literal["excel", "image", "video"], db: Session):
    ext_map = {
        "excel": ([".xlsx", ".xls"], ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet","application/vnd.ms-excel"]),
        "image": ([".png", ".jpg", ".jpeg"], ["image/png", "image/jpeg"]),
        "video": ([".mp4", ".avi", ".mov", ".mkv"], ["video/mp4", "video/x-msvideo", "video/quicktime", "video/x-matroska"])
    }

    allowed_exts, allowed_mimes = ext_map[file_type]
    filename = file.filename.lower()
    ext = os.path.splitext(filename)[1]

    if ext not in allowed_exts:
        raise HTTPException(status_code=400, detail=f"File extension {ext} not allowed for type {file_type}")
    
    if file.content_type not in allowed_mimes and not file.content_type.startswith("video/"):
        raise HTTPException(status_code=400, detail=f"MIME type {file.content_type} is invalid for {file_type}")
    
    content = await file.read()
    size_kb = len(content) / 1024
    await file.seek(0)

    # Simpan file ke folder
    save_path = os.path.join(UPLOAD_DIRS[file_type], filename)
    with open(save_path, "wb") as out_file: # context manager open file dalam mode biner (wb - write biner), out_file object file tujuan
        shutil.copyfileobj(file.file, out_file) # copy file.file ke out_file

    # Simpan metadata ke database
    metadata = FileMetadata(
        filename=filename,
        file_type=file_type,
        mime_type=file.content_type,
        size_kb=round(size_kb, 2),
        saved_path=save_path
    )

    db.add(metadata)
    db.commit()
    db.refresh(metadata)

    response = {
        "id": metadata.id,
        "filename": filename,
        "type": file_type,
        "mime": file.content_type,
        "size_kb": metadata.size_kb,
        "saved_to": save_path
    }

    # Jika Excel, tampilkan isi sheet pertama
    if file_type == "excel":
        wb = load_workbook(save_path)
        ws = wb.active
        data = [[cell.value for cell in row] for row in ws.iter_rows(min_row=1, max_row=5)]
        response["preview_data"] = data

    return response